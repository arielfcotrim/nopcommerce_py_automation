import openpyxl


def get_row_count(file, sheet_name):
    """
    Get the number of rows in a worksheet.

    :param file: The Excel file containing the worksheet.
    :param sheet_name: The name of the worksheet.
    :return: The number of rows in the worksheet.
    """
    # Load the workbook and worksheet
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook[sheet_name]

    # Return the number of rows in the worksheet
    return worksheet.max_row


def get_column_count(file, sheet_name):
    """
    Get the number of columns in a worksheet.

    :param file: The Excel file containing the worksheet.
    :param sheet_name: The name of the worksheet.
    :return: The number of columns in the worksheet.
    """
    # Load the workbook and worksheet
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook[sheet_name]

    # Return the number of columns in the worksheet
    return worksheet.max_column


def read_data(file, sheet_name, row_number, column_number):
    """
    Read data from a worksheet cell.

    :param file: The Excel file containing the worksheet.
    :param sheet_name: The name of the worksheet.
    :param row_number: The row number of the cell.
    :param column_number: The column number of the cell.
    :return: The data in the specified cell.
    """
    # Load the workbook and worksheet
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook[sheet_name]

    # Return the data in the specified cell
    return worksheet.cell(row=row_number, column=column_number).value


def write_data(file, sheet_name, row_number, column_number, data):
    """
    Write data to a worksheet cell.

    :param file: The Excel file containing the worksheet.
    :param sheet_name: The name of the worksheet.
    :param row_number: The row number of the cell.
    :param column_number: The column number of the cell.
    :param data: The data to be written to the cell.
    :return: None
    """
    # Load the workbook and worksheet
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook[sheet_name]

    # Write the data to the specified cell
    worksheet.cell(row=row_number, column=column_number).value = data

    # Save the workbook
    workbook.save(file)
